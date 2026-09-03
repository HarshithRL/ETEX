from __future__ import annotations

from pathlib import Path

from ai_brain.core.procurement_ai.insights import build_insight_payload
from ai_brain.core.procurement_ai.packs.compare_xlsx import build_comparison_xlsx
from ai_brain.core.procurement_ai.packs.steerco_ppt import build_steerco_ppt
from ai_brain.core.procurement_ai.process_type import classify_process_type


class _Project:
    id = "proj-demo"
    code = "PR-100"
    name = "SWIFT CSP assessment"
    business_process = "Indirect"
    category = "SWIFT CSCF"
    description = "Professional services days"


class _Artifact:
    id = "a1"
    original_name = "EY_proposal.pdf"
    parse_status = "ok"


def test_classify_swift_as_it_services():
    assert classify_process_type(
        business_process="Indirect",
        category="SWIFT CSCF",
        name="SWIFT CSP",
    ) == "indirect_it_services"


def test_classify_direct_screws():
    assert classify_process_type(business_process="Direct", name="TG screws") == "direct_tg"


def test_vendor_names_from_swift_filenames():
    class _KPMG:
        id = "k1"
        original_name = "KPMG_Etex_SWIFT CSP_v20260313.pdf"
        parse_status = "ok"

    class _RFP:
        id = "r1"
        original_name = "Etex_RFP_Swift CSP_06022026.pdf"
        parse_status = "ok"

    class _TF:
        id = "t1"
        original_name = "TF Inetum Proposal Etex CSP Assessment.pdf"
        parse_status = "ok"

    class _Untitled:
        id = "proj-demo"
        code = "PR-100"
        name = "untitled"
        business_process = "untitled"
        category = "untitled"
        description = "untitled"

    class _EY:
        id = "e1"
        original_name = "EY - ETEX Group - SWIFT Customer Security Program (CSP) independent assessment proposal 2026 (vF).pdf"
        parse_status = "ok"

    insights = build_insight_payload(_Untitled(), [_KPMG(), _RFP(), _TF(), _EY()], [])
    assert insights["process_type"] == "indirect_it_services"
    names = [vendor["name"] for vendor in insights["vendors"]]
    assert names == ["KPMG", "TriFinance + Inetum", "EY"]


def test_pack_builders_write_files(tmp_path, monkeypatch):
    monkeypatch.setenv("MATE_PROJECTS_DATA_ROOT", str(tmp_path))
    insights = build_insight_payload(_Project(), [_Artifact()], [])
    assert insights["process_type"] == "indirect_it_services"
    assert insights["decision"]["can_build_xlsx"] is True
    xlsx = build_comparison_xlsx("proj-demo", insights, thread_id="t1")
    assert xlsx["status"] == "ready"
    assert xlsx["llm_used"] is False
    assert Path(xlsx["path"]).exists()
    from openpyxl import load_workbook

    sheets = load_workbook(xlsx["path"]).sheetnames
    assert "Final BAFO Comparison" in sheets
    assert "Procurement Requirements" in sheets
    assert "Sources & Notes" in sheets
    assert "Cover" in sheets
    bafo = load_workbook(xlsx["path"])["Final BAFO Comparison"]
    assert bafo["A3"].value == "Solution costs (Software, Materials)"
    assert bafo["B3"].value == "Recurring"
    assert bafo["A17"].value == "Score Card score"
    reqs = load_workbook(xlsx["path"])["Procurement Requirements"]
    assert reqs["A5"].value == "R01"
    assert reqs["A44"].value == "R40"
    ppt = build_steerco_ppt("proj-demo", thread_id="t1")
    assert ppt["status"] == "ready"
    assert Path(ppt["path"]).exists()


def test_direct_pack_uses_france_sheets(tmp_path, monkeypatch):
    monkeypatch.setenv("MATE_PROJECTS_DATA_ROOT", str(tmp_path))

    class _Direct:
        id = "proj-direct"
        code = "PR-200"
        name = "TG screws France"
        business_process = "Direct"
        category = "fasteners"
        description = "P×Q basket"

    insights = build_insight_payload(_Direct(), [_Artifact()], [])
    assert insights["process_type"] == "direct_tg"
    xlsx = build_comparison_xlsx("proj-direct", insights, thread_id="t2")
    from openpyxl import load_workbook

    sheets = load_workbook(xlsx["path"]).sheetnames
    assert "extra-polation " in sheets
    assert "Slide 2 - Supplier Ranking" in sheets
    assert "Samples " in sheets
    assert "Slide 5 - Commercial Terms" in sheets
    assert "Slide 7 - RACI" in sheets

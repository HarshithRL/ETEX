"""Workbook chrome shared by Direct and Indirect packs."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ai_brain.core.procurement_ai.packs.matrix_layout import MISSING

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color="FF6900")
WRAP = Alignment(wrap_text=True, vertical="center")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
MISSING_FILL = PatternFill("solid", fgColor="FFF2CC")

COVER_FIELDS = (
    ("FIELD_PROCESS_TYPE", "process_type"),
    ("FIELD_PROCESS_LABEL", "process_label"),
    ("FIELD_OWNER_ENTITY", "owner_entity"),
    ("FIELD_PROJECT_CODE", "project_code"),
    ("FIELD_PROJECT_NAME", "project_name"),
    ("FIELD_KNOWLEDGE_PCT", "knowledge_pct"),
    ("FIELD_KB_STATUS", "kb_status"),
)


def header_row(ws: Worksheet, headers: list[str], row: int = 1, start_col: int = 1) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def put(ws: Worksheet, row: int, col: int, value: Any, *, wrap: bool = True) -> None:
    text = MISSING if value is None or value == "" else value
    cell = ws.cell(row, col, text)
    if wrap:
        cell.alignment = WRAP
    cell.border = THIN
    if str(text).strip().lower() == MISSING:
        cell.fill = MISSING_FILL


def autosize(ws: Worksheet, max_width: int = 42) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "").split("\n")[0]) for cell in column)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), max_width)


def vendor_names(insights: dict[str, Any], extracted_vendors: list[Any] | None = None) -> list[str]:
    names: list[str] = []
    if extracted_vendors:
        names = [
            getattr(v, "name", None) or (v.get("name") if isinstance(v, dict) else None) or ""
            for v in extracted_vendors
        ]
    names = [name.strip() for name in names if name and str(name).strip()]
    if not names:
        names = [v.get("name") or "Unnamed vendor" for v in insights.get("vendors") or []]
        names = [name.strip() for name in names if name and str(name).strip()]
    return names or ["Unnamed vendor"]


def col_letter(index: int) -> str:
    return get_column_letter(index)

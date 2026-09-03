"""Deterministic comparison workbook. Excel is the source of truth."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ai_brain.core.procurement_ai.extraction import load_facts, overlay_insights, scan_chunks_for_facts
from ai_brain.core.procurement_ai.packs import store
from ai_brain.core.procurement_ai.project_context import load_context

COVER_FIELDS = (
    ("FIELD_PROCESS_TYPE", "process_type"),
    ("FIELD_PROCESS_LABEL", "process_label"),
    ("FIELD_OWNER_ENTITY", "owner_entity"),
    ("FIELD_PROJECT_CODE", "project_code"),
    ("FIELD_PROJECT_NAME", "project_name"),
    ("FIELD_KNOWLEDGE_PCT", "knowledge_pct"),
    ("FIELD_KB_STATUS", "kb_status"),
)


def build_comparison_xlsx(project_id: str, insights: dict[str, Any], *, thread_id: str = "") -> dict[str, Any]:
    store.write_status(project_id, "xlsx", status="running", thread_id=thread_id or None)
    chunks: list[Any] = []
    try:
        _project, _artifacts, chunks = load_context(project_id)
    except Exception:  # noqa: BLE001
        chunks = []
    facts = load_facts(project_id) or scan_chunks_for_facts(insights, chunks)
    insights = overlay_insights(insights, facts)
    filename = "comparison_matrix.xlsx"
    path = store.pack_file(project_id, filename)
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    _write_cover(cover, insights)
    _write_vendors(wb.create_sheet("Vendors"), insights)
    _write_requirements(wb.create_sheet("Requirements"), insights)
    _write_scorecard(wb.create_sheet("Scorecard"), insights)
    _write_tco(wb.create_sheet("TCO"), insights)
    _write_red_flags(wb.create_sheet("RedFlags"), insights)
    _write_field_map(wb.create_sheet("FieldMap"), insights)
    wb.save(path)
    href = store.flask_href(project_id, "xlsx")
    store.write_status(project_id, "xlsx", status="ready", href=href, filename=filename, thread_id=thread_id or None, error=None)
    return {"status": "ready", "href": href, "filename": filename, "path": str(path)}


def _write_cover(ws, insights: dict[str, Any]) -> None:
    ws["A1"] = "Nexus Mate comparison pack"
    ws["A1"].font = Font(bold=True, size=16, color="FF6900")
    ws["A2"] = "Draft workbook. Humans award. Empty is missing, never zero."
    ws["A4"] = "Named field"
    ws["B4"] = "Value"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)
    for row, (field, key) in enumerate(COVER_FIELDS, start=5):
        ws[f"A{row}"] = field
        ws[f"B{row}"] = insights.get(key, "")
    ws["A12"] = "FIELD_DECISION_SUMMARY"
    ws["B12"] = (insights.get("decision") or {}).get("summary") or ""
    ws["A13"] = "FIELD_WEIGHTS_LOCKED"
    ws["B13"] = "NO — draft scores only"
    ws["A14"] = "FIELD_AWARD"
    ws["B14"] = "Not awarded"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72


def _write_vendors(ws, insights: dict[str, Any]) -> None:
    _header_row(ws, ["Vendor", "Role", "Files", "Coverage %", "Headline", "Evidence locator"])
    for index, vendor in enumerate(insights.get("vendors") or [], start=2):
        evidence = vendor.get("evidence") or []
        ws[f"A{index}"] = vendor.get("name")
        ws[f"B{index}"] = vendor.get("role")
        ws[f"C{index}"] = vendor.get("artifact_count")
        ws[f"D{index}"] = vendor.get("coverage_pct")
        ws[f"E{index}"] = vendor.get("headline")
        ws[f"F{index}"] = evidence[0].get("locator") if evidence else "missing"
    _autosize(ws)


def _write_requirements(ws, insights: dict[str, Any]) -> None:
    _header_row(ws, ["Checklist key", "Label", "Severity", "Status", "Needed"])
    for index, item in enumerate((insights.get("requirements") or {}).get("items") or [], start=2):
        ws[f"A{index}"] = item.get("checklist_key")
        ws[f"B{index}"] = item.get("label")
        ws[f"C{index}"] = item.get("severity")
        ws[f"D{index}"] = item.get("status")
        ws[f"E{index}"] = item.get("needed") or ""
    _autosize(ws)


def _write_tco(ws, insights: dict[str, Any]) -> None:
    process = insights.get("process_type") or ""
    _header_row(ws, ["Vendor", "External cost", "Internal days", "Day rate", "Internal €800/day", "TCO", "Source"])
    note = "P×Q — quote missing, do not write 0" if process == "direct_tg" else "Licence + impl or professional-services days — quote missing, do not write 0"
    vendors = insights.get("vendors") or []
    if not vendors:
        ws["A2"] = "No vendor column yet"
        ws["F2"] = "missing"
        return
    for index, vendor in enumerate(vendors, start=2):
        evidence = vendor.get("evidence") or []
        locator = evidence[0].get("locator") if evidence else "missing"
        ws[f"A{index}"] = vendor.get("name")
        ws[f"B{index}"] = vendor.get("external_cost") or "missing"
        ws[f"C{index}"] = vendor.get("internal_days") or "missing"
        ws[f"D{index}"] = vendor.get("day_rate") or "missing"
        ws[f"E{index}"] = 800
        ws[f"F{index}"] = "missing"
        ws[f"G{index}"] = locator if locator != "missing" else note
    _autosize(ws)


def _write_scorecard(ws, insights: dict[str, Any]) -> None:
    extracted = (insights.get("requirements") or {}).get("extracted") or []
    vendors = [v.get("name") for v in insights.get("vendors") or [] if v.get("name")]
    headers = ["Req id", "Label", "Severity", *vendors, "Evidence"]
    _header_row(ws, headers)
    rows = extracted or (insights.get("requirements") or {}).get("items") or []
    if not rows:
        ws["A2"] = "No requirements extracted"
        ws["B2"] = "missing"
        return
    for index, item in enumerate(rows, start=2):
        ws[f"A{index}"] = item.get("id") or item.get("checklist_key")
        ws[f"B{index}"] = item.get("label")
        ws[f"C{index}"] = item.get("severity") or ""
        status_map = item.get("vendor_status") or {}
        for col, vendor in enumerate(vendors, start=4):
            ws.cell(index, col, status_map.get(vendor) or item.get("status") or "missing")
        evidence = item.get("evidence") or []
        quote = evidence[0].get("locator") if evidence else item.get("needed") or ""
        ws.cell(index, 4 + len(vendors), quote)
    _autosize(ws)


def _write_red_flags(ws, insights: dict[str, Any]) -> None:
    _header_row(ws, ["Severity", "Item"])
    blockers = (insights.get("decision") or {}).get("blockers") or []
    if not blockers:
        ws["A2"] = "info"
        ws["B2"] = "No blocking gaps recorded from filenames. Confirm scope in SteerCo."
        return
    for index, item in enumerate(blockers, start=2):
        ws[f"A{index}"] = "blocking"
        ws[f"B{index}"] = item
    _autosize(ws)


def _write_field_map(ws, insights: dict[str, Any]) -> None:
    _header_row(ws, ["PPT field", "Excel sheet", "Cell / column"])
    rows = [
        ("FIELD_PROCESS_LABEL", "Cover", "B6"),
        ("FIELD_OWNER_ENTITY", "Cover", "B7"),
        ("FIELD_PROJECT_CODE", "Cover", "B8"),
        ("FIELD_PROJECT_NAME", "Cover", "B9"),
        ("FIELD_KNOWLEDGE_PCT", "Cover", "B10"),
        ("FIELD_DECISION_SUMMARY", "Cover", "B12"),
        ("FIELD_WEIGHTS_LOCKED", "Cover", "B13"),
        ("FIELD_AWARD", "Cover", "B14"),
        ("FIELD_VENDOR_NAME", "Vendors", "A"),
        ("FIELD_VENDOR_HEADLINE", "Vendors", "E"),
        ("FIELD_TCO", "TCO", "F"),
    ]
    for index, row in enumerate(rows, start=2):
        ws[f"A{index}"] = row[0]
        ws[f"B{index}"] = row[1]
        ws[f"C{index}"] = row[2]
    _autosize(ws)


def _header_row(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="202020")
    font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")


def _autosize(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 60)


def existing_xlsx_path(project_id: str) -> Path | None:
    status = store.read_status(project_id).get("xlsx") or {}
    filename = status.get("filename")
    if not filename:
        return None
    path = store.pack_file(project_id, filename)
    return path if path.exists() else None

"""Direct workbook: France overview P×Q / ranking / samples / terms / RACI."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ai_brain.core.procurement_ai.packs.matrix_layout import (
    MISSING,
    RACI_ACTIVITIES,
    RACI_DEFAULT,
    RACI_ROLES,
)
from ai_brain.core.procurement_ai.packs.matrix_schema import DirectMatrix, DirectVendorTerms
from ai_brain.core.procurement_ai.packs.xlsx_style import autosize, header_row, put, vendor_names


def write_direct(wb, insights: dict[str, Any], matrix: DirectMatrix) -> None:
    vendors = _vendors(insights, matrix)
    _write_start(wb.create_sheet("START HERE - Management Slides"), insights, matrix)
    _write_pq(wb.create_sheet("extra-polation "), vendors, matrix)
    _write_ranking(wb.create_sheet("Slide 2 - Supplier Ranking"), vendors, matrix)
    _write_samples(wb.create_sheet("Samples "), vendors)
    _write_terms(wb.create_sheet("Slide 5 - Commercial Terms"), vendors, matrix)
    _write_raci(wb.create_sheet("Slide 7 - RACI"))


def _vendors(insights: dict[str, Any], matrix: DirectMatrix) -> list[DirectVendorTerms]:
    by_name = {v.name: v for v in matrix.vendors if v.name}
    ordered: list[DirectVendorTerms] = []
    for name in vendor_names(insights, matrix.vendors):
        ordered.append(by_name.get(name) or DirectVendorTerms(name=name))
    return ordered or [DirectVendorTerms(name="Unnamed vendor")]


def _write_start(ws: Worksheet, insights: dict[str, Any], matrix: DirectMatrix) -> None:
    ws["B2"] = f"{insights.get('project_name') or 'Supplier selection'} — Management Slideshow"
    ws["B2"].font = Font(bold=True, size=16, color="FF6900")
    ws["B3"] = (
        "Open the slide tabs in order. Draft pack from uploaded files. "
        "Empty is missing, never zero. Humans award."
    )
    ws["B3"].alignment = Alignment(wrap_text=True)
    header_row(ws, ["#", "Slide", "Purpose"], row=15, start_col=2)
    agenda = (
        (1, "Executive Summary", "Overall recommendation view and decision message"),
        (2, "Supplier Ranking", "Cost ranking from the P×Q basket"),
        (3, "3Y Cost View", "Cost development across volume years"),
        (4, "Samples Readiness", "Qualification and sample follow-ups"),
        (5, "Commercial Terms", "Non-price terms: agreements, payment, delivery, rebates"),
        (6, "Next Steps", "Workplan to finalize supplier selection"),
        (7, "RACI", "Ownership and decision governance"),
    )
    for offset, (num, slide, purpose) in enumerate(agenda, start=16):
        ws.cell(offset, 2, num)
        ws.cell(offset, 3, slide)
        ws.cell(offset, 4, purpose)
    ws["B24"] = matrix.summary or (insights.get("decision") or {}).get("summary") or ""
    ws["B24"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 56


def _write_pq(ws: Worksheet, vendors: list[DirectVendorTerms], matrix: DirectMatrix) -> None:
    from ai_brain.core.procurement_ai.packs.xlsx_style import HEADER_FILL, HEADER_FONT, WRAP

    ws["B2"] = 2023
    ws["C2"] = 2024
    ws["D2"] = 2025
    ws["A3"] = "SKU / article"
    ws["B3"] = "Volumes"
    ws["C3"] = "Volumes"
    ws["D3"] = "Volumes"
    for index, vendor in enumerate(vendors):
        price_cell = ws.cell(3, 5 + index, vendor.name)
        price_cell.fill = HEADER_FILL
        price_cell.font = HEADER_FONT
        price_cell.alignment = WRAP
        cost_cell = ws.cell(2, 5 + len(vendors) + index, f"Cost ({vendor.name})")
        cost_cell.alignment = WRAP
    for col in range(1, 5):
        cell = ws.cell(3, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    lines = matrix.basket_lines or []
    if not lines:
        ws["A4"] = MISSING
        for col in range(2, 5 + len(vendors)):
            put(ws, 4, col, MISSING)
        autosize(ws)
        return
    for row_offset, line in enumerate(lines, start=4):
        sku = line.sku or line.description or MISSING
        ws.cell(row_offset, 1, sku)
        for col, qty in enumerate((line.qty_2023, line.qty_2024, line.qty_2025), start=2):
            put(ws, row_offset, col, qty)
        for index, vendor in enumerate(vendors):
            price = (line.unit_prices or {}).get(vendor.name) or MISSING
            put(ws, row_offset, 5 + index, price)
            cost_col = 5 + len(vendors) + index
            if _numeric(line.qty_2023) and _numeric(price):
                vol_letter = get_column_letter(2)
                price_letter = get_column_letter(5 + index)
                ws.cell(
                    row_offset,
                    cost_col,
                    f"={vol_letter}{row_offset}*{price_letter}{row_offset}",
                )
            else:
                put(ws, row_offset, cost_col, MISSING)
    autosize(ws, max_width=22)
    ws.freeze_panes = "E4"


def _write_ranking(ws: Worksheet, vendors: list[DirectVendorTerms], matrix: DirectMatrix) -> None:
    ws["B2"] = "Supplier Ranking — Selected suppliers"
    ws["B2"].font = Font(bold=True, size=16)
    ws["B3"] = "Draft ranking from quoted P×Q. Scores stay draft. Humans award."
    header_row(
        ws,
        ["Rank", "Supplier", "Selected basket", "Total Cost 2023", "Total Cost 2024", "Total Cost 2025", "Key note"],
        row=6,
        start_col=2,
    )
    notes = list(matrix.ranking_notes or [])
    for index, vendor in enumerate(vendors, start=1):
        row = 6 + index
        ws.cell(row, 2, index)
        ws.cell(row, 3, vendor.name)
        for col in range(4, 8):
            put(ws, row, col, MISSING)
        put(ws, row, 8, notes[index - 1] if index - 1 < len(notes) else vendor.note or MISSING)
    autosize(ws, max_width=28)


def _write_samples(ws: Worksheet, vendors: list[DirectVendorTerms]) -> None:
    header_row(
        ws,
        ["Supplier", "Launch Ariba", "Qualified", "RFI offer", "RFP offer", "Steerco", "Samples"],
    )
    for index, vendor in enumerate(vendors, start=2):
        put(ws, index, 1, vendor.name)
        put(ws, index, 2, MISSING)
        put(ws, index, 3, vendor.qualified)
        put(ws, index, 4, vendor.rfi_offer)
        put(ws, index, 5, vendor.rfp_offer)
        put(ws, index, 6, MISSING)
        put(ws, index, 7, vendor.samples)
    autosize(ws)


def _write_terms(ws: Worksheet, vendors: list[DirectVendorTerms], matrix: DirectMatrix) -> None:
    ws["B2"] = "Commercial Terms Overview"
    ws["B2"].font = Font(bold=True, size=16)
    header_row(
        ws,
        [
            "Rank",
            "Supplier",
            "Agreement / position",
            "Delivery / incoterm",
            "Lead time",
            "Payment terms",
            "Rebate / indexation",
            "Management view",
            "Follow-up",
        ],
        row=6,
        start_col=2,
    )
    for index, vendor in enumerate(vendors, start=1):
        row = 6 + index
        ws.cell(row, 2, index)
        ws.cell(row, 3, vendor.name)
        put(ws, row, 4, vendor.agreement)
        put(ws, row, 5, vendor.incoterm)
        put(ws, row, 6, vendor.lead_time)
        put(ws, row, 7, vendor.payment_terms)
        put(ws, row, 8, vendor.rebate)
        put(ws, row, 9, vendor.note or MISSING)
        put(ws, row, 10, vendor.follow_up)
    takeaway_row = 8 + len(vendors)
    ws.cell(
        takeaway_row,
        2,
        matrix.summary or "Keep pricing separate from terms view. Humans award.",
    )
    autosize(ws, max_width=32)


def _write_raci(ws: Worksheet) -> None:
    ws["B2"] = "RACI Overview"
    ws["B2"].font = Font(bold=True, size=16)
    ws["B3"] = "Governance for final validation and supplier award."
    header_row(ws, ["Activity", *RACI_ROLES], row=6, start_col=2)
    for offset, (activity, marks) in enumerate(zip(RACI_ACTIVITIES, RACI_DEFAULT, strict=True), start=7):
        ws.cell(offset, 2, activity)
        for col, mark in enumerate(marks, start=3):
            ws.cell(offset, col, mark)
    ws.cell(16, 2, "RACI legend: R = Responsible | A = Accountable | C = Consulted | I = Informed")
    ws.cell(
        21,
        2,
        "Decision principle: Procurement drives the recommendation, functional owners validate, Steerco awards.",
    )
    autosize(ws, max_width=28)
    ws.column_dimensions["B"].width = 40


def _numeric(value: Any) -> bool:
    if value is None or value == MISSING:
        return False
    try:
        float(str(value).replace(",", "").replace("€", "").strip())
        return True
    except ValueError:
        return False

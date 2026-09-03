"""Indirect workbook: SWIFT CSP Vendor Comparison.xlsx layout."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ai_brain.core.procurement_ai.packs.matrix_layout import (
    BAFO_ATTRS,
    BAFO_ROWS,
    MISSING,
    SCORED,
    SCORED_BY_ETEX,
    SWIFT_REQUIREMENTS,
)
from ai_brain.core.procurement_ai.packs.matrix_schema import IndirectMatrix, IndirectVendorFacts, blank_indirect_vendor
from ai_brain.core.procurement_ai.packs.xlsx_style import autosize, header_row, put, vendor_names


def write_indirect(wb, insights: dict[str, Any], matrix: IndirectMatrix) -> None:
    vendors = _vendors(insights, matrix)
    _write_bafo(wb.create_sheet("Final BAFO Comparison"), vendors)
    _write_requirements(wb.create_sheet("Procurement Requirements"), insights, matrix, vendors)
    _write_sources(wb.create_sheet("Sources & Notes"), insights, matrix, vendors)


def _vendors(insights: dict[str, Any], matrix: IndirectMatrix) -> list[IndirectVendorFacts]:
    extracted = {_norm_name(v.name): v for v in matrix.vendors if v.name}
    ordered: list[IndirectVendorFacts] = []
    used: set[str] = set()
    for name in vendor_names(insights, None):
        key = _norm_name(name)
        match = extracted.get(key) or _fuzzy_vendor(extracted, name)
        if match is not None:
            used.add(_norm_name(match.name))
            used.add(key)
            ordered.append(match.model_copy(update={"name": name}))
        else:
            ordered.append(blank_indirect_vendor(name))
    for vendor in matrix.vendors:
        if vendor.name and _norm_name(vendor.name) not in used:
            ordered.append(vendor)
    return ordered or [blank_indirect_vendor("Unnamed vendor")]


def _norm_name(name: str) -> str:
    return " ".join(name.lower().replace("+", " ").split())


def _fuzzy_vendor(extracted: dict[str, IndirectVendorFacts], name: str) -> IndirectVendorFacts | None:
    needle = _norm_name(name)
    for key, vendor in extracted.items():
        if needle in key or key in needle:
            return vendor
    return None


def _write_bafo(ws: Worksheet, vendors: list[IndirectVendorFacts]) -> None:
    ws["B1"] = "Type "
    ws["B1"].font = Font(bold=True)
    for index, vendor in enumerate(vendors):
        col = 3 + index
        ws.cell(1, col, vendor.name).font = Font(bold=True)
        ws.cell(2, col, vendor.legal_name or f"Vendor {index + 1} - {vendor.name}")
        ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 36
    for row_offset, ((label, typ), attr) in enumerate(zip(BAFO_ROWS, BAFO_ATTRS, strict=True), start=3):
        ws.cell(row_offset, 1, label)
        ws.cell(row_offset, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row_offset, 2, typ)
        ws.cell(row_offset, 2).alignment = Alignment(wrap_text=True, vertical="center")
        if not label:
            continue
        for index, vendor in enumerate(vendors):
            raw = getattr(vendor, attr, MISSING) if attr else ""
            if label in SCORED_BY_ETEX and (not raw or raw == MISSING):
                raw = SCORED
            put(ws, row_offset, 3 + index, raw or MISSING)
        ws.row_dimensions[row_offset].height = 32 if len(label) > 40 else 22
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 16
    for index in range(len(vendors)):
        ws.column_dimensions[get_column_letter(3 + index)].width = 28
    ws.freeze_panes = "C3"


def _write_requirements(
    ws: Worksheet,
    insights: dict[str, Any],
    matrix: IndirectMatrix,
    vendors: list[IndirectVendorFacts],
) -> None:
    title = (
        f"Procurement Requirements Compliance Matrix - "
        f"{insights.get('project_name') or 'Etex'} "
        f"({insights.get('project_code') or ''})".strip()
    )
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = (
        "Legend:  ✓ = compliant / evidenced    ~ = partial / conditional    "
        "✗ = gap or deviation    ? = not evidenced in submission    — = Etex-side process step"
    )
    headers = ["ID", "Requirement", "Category", "Target / Detail", *[v.name for v in vendors]]
    header_row(ws, headers, row=4)
    marks_by_id = {item.requirement_id: item.vendor_marks for item in matrix.requirements}
    for offset, (rid, label, category, target) in enumerate(SWIFT_REQUIREMENTS, start=5):
        ws.cell(offset, 1, rid)
        ws.cell(offset, 2, label)
        ws.cell(offset, 3, category)
        ws.cell(offset, 4, target)
        vendor_marks = marks_by_id.get(rid) or {}
        for index, vendor in enumerate(vendors):
            put(ws, offset, 5 + index, vendor_marks.get(vendor.name) or MISSING)
    autosize(ws, max_width=36)
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["D"].width = 36
    ws.freeze_panes = "E5"


def _write_sources(
    ws: Worksheet,
    insights: dict[str, Any],
    matrix: IndirectMatrix,
    vendors: list[IndirectVendorFacts],
) -> None:
    ws["A1"] = "Data sources (parsed project uploads)"
    ws["A1"].font = Font(bold=True)
    row = 3
    for index, vendor in enumerate(vendors, start=1):
        ws.cell(row, 1, f"Vendor {index} - {vendor.legal_name or vendor.name}")
        files = vendor.source_files
        if not files:
            insight = next((v for v in insights.get("vendors") or [] if v.get("name") == vendor.name), None)
            files = (insight or {}).get("headline") or MISSING
        put(ws, row, 2, files)
        row += 1
    ws.cell(row, 1, "RFP")
    put(ws, row, 2, "See uploaded RFP / requirements files")
    row += 2
    ws.cell(row, 1, "Method notes")
    put(
        ws,
        row,
        2,
        matrix.method_notes
        or "Cost rows use each vendor's quoted figures when present. "
        "Etex man-days valued at EUR 800/day. Empty is missing, never zero. Humans award.",
    )
    row += 1
    ws.cell(row, 1, "Key flags")
    flags = matrix.flags or (insights.get("decision") or {}).get("blockers") or []
    put(ws, row, 2, "; ".join(str(item) for item in flags) or MISSING)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 80

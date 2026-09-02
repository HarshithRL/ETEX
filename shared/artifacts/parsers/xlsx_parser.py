from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from ..blocks import make_content_block
from ..exceptions import ArtifactParseError, CorruptArtifact
from ..ids import file_artifact_id
from ..models import (
    ArtifactDocument,
    ArtifactMetadata,
    ArtifactType,
    BlockType,
    ContentBlock,
    MIME_BY_TYPE,
    ParseOptions,
    SourceLocation,
)
from .pdf_tables import drop_empty_columns, table_to_text


class XlsxParser:
    kind = ArtifactType.EXCEL

    def parse(self, path: Path, options: ParseOptions) -> ArtifactDocument:
        try:
            workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
        except InvalidFileException as exc:
            raise CorruptArtifact(f"Unreadable Excel file: {path.name}") from exc
        except Exception as exc:
            raise CorruptArtifact(f"Failed to open Excel file: {path.name}") from exc

        formula_book = None
        try:
            formula_book = load_workbook(filename=str(path), data_only=False, read_only=True)
        except Exception:
            formula_book = None

        props = workbook.properties
        blocks: list[ContentBlock] = []
        warnings: list[str] = []
        index = 0
        visible_sheets = 0

        try:
            for sheet in workbook.worksheets:
                hidden = sheet.sheet_state != "visible"
                if hidden and not options.include_hidden_sheets:
                    continue
                visible_sheets += 1
                rows, formulas, cell_range = _extract_sheet(sheet, formula_book)
                if not rows:
                    warnings.append(f'Sheet "{sheet.title}" has no used cells.')
                    continue
                if not options.include_tables:
                    text = "\n".join(" | ".join(row) for row in rows)
                    index += 1
                    blocks.append(
                        make_content_block(
                            seq_id=f"excel-{index:04d}",
                            block_type=BlockType.TEXT,
                            text=text,
                            location=SourceLocation(sheet=sheet.title, cell_range=cell_range),
                        )
                    )
                    continue
                index += 1
                extra = {"hidden": hidden}
                if formulas:
                    extra["formulas"] = formulas
                blocks.append(
                    make_content_block(
                        seq_id=f"excel-{index:04d}",
                        block_type=BlockType.TABLE,
                        text=table_to_text(rows),
                        table=rows,
                        location=SourceLocation(sheet=sheet.title, cell_range=cell_range),
                        extra=extra,
                    )
                )
        except Exception as exc:
            raise ArtifactParseError(f"Failed to extract Excel {path.name}: {exc}") from exc
        finally:
            workbook.close()
            if formula_book is not None:
                formula_book.close()

        if not blocks:
            warnings.append("Excel workbook contained no extractable cell values.")

        return ArtifactDocument(
            source=str(path),
            artifact_type=ArtifactType.EXCEL,
            artifact_id=file_artifact_id(path),
            metadata=ArtifactMetadata(
                filename=path.name,
                artifact_type=ArtifactType.EXCEL,
                mime_type=MIME_BY_TYPE[ArtifactType.EXCEL],
                title=_meta_str(props.title),
                author=_meta_str(props.creator),
                created=_meta_str(props.created),
                modified=_meta_str(props.modified),
                sheet_count=visible_sheets,
            ),
            blocks=blocks,
            warnings=warnings,
        )


def _extract_sheet(sheet, formula_book) -> tuple[list[list[str]], dict[str, str], str | None]:
    rows: list[list[str]] = []
    formulas: dict[str, str] = {}
    min_row = sheet.min_row
    max_row = sheet.max_row
    min_col = sheet.min_column
    max_col = sheet.max_column
    if min_row is None or max_row is None or min_col is None or max_col is None:
        return [], {}, None

    formula_sheet = None
    if formula_book is not None and sheet.title in formula_book.sheetnames:
        formula_sheet = formula_book[sheet.title]

    for r in range(min_row, max_row + 1):
        row_vals: list[str] = []
        for c in range(min_col, max_col + 1):
            value = sheet.cell(r, c).value
            raw = formula_sheet.cell(r, c).value if formula_sheet is not None else None
            if isinstance(raw, str) and raw.startswith("="):
                formulas[f"{get_column_letter(c)}{r}"] = raw
            if value is None and isinstance(raw, str):
                cell_text = raw
            elif value is None:
                cell_text = ""
            else:
                cell_text = str(value).strip()
            row_vals.append(cell_text)
        if any(row_vals):
            rows.append(row_vals)

    rows = drop_empty_columns(rows)
    cell_range = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )
    return rows, formulas, cell_range


def _meta_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None

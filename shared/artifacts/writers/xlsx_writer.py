from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..exceptions import ArtifactPatchError, ArtifactWriteError
from ..models import ArtifactDocument, ArtifactType
from ..patch import LocationTracker, cell_value, parse_cell_ref, require_table, require_text
from ..spec import ArtifactSpec, PatchOp, SpecSheet


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_BODY_FONT = Font(name="Calibri", size=11)


class XlsxWriter:
    kind = ArtifactType.EXCEL

    def write(self, spec: ArtifactSpec, dest: Path) -> None:
        try:
            workbook = Workbook()
            default = workbook.active
            sheets = list(spec.sheets or [SpecSheet(name="Sheet1")])
            for index, sheet in enumerate(sheets):
                if index == 0:
                    ws = default
                    ws.title = _sheet_title(sheet.name, index)
                else:
                    ws = workbook.create_sheet(title=_sheet_title(sheet.name, index))
                _write_sheet(ws, sheet)
                if sheet.hidden:
                    ws.sheet_state = "hidden"
            if spec.title:
                workbook.properties.title = spec.title
            if spec.author:
                workbook.properties.creator = spec.author
            dest.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(str(dest))
        except ArtifactWriteError:
            raise
        except Exception as exc:
            raise ArtifactWriteError(f"Failed to write Excel file {dest.name}: {exc}") from exc

    def apply_ops(
        self,
        path: Path,
        document: ArtifactDocument,
        ops: list[PatchOp],
        dest: Path,
    ) -> None:
        workbook = load_workbook(filename=str(path), data_only=False)
        tracker = LocationTracker(document.blocks)
        for op in ops:
            _apply_xlsx_op(workbook, tracker, op)
        dest.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(dest))


def _write_sheet(ws: Worksheet, sheet: SpecSheet) -> None:
    rows = [list(row) for row in sheet.rows]
    _write_rows(ws, rows)
    for coord, value in (sheet.cells or {}).items():
        ws[str(coord)] = value
    _autosize(ws, rows)


def _write_rows(ws: Worksheet, rows: list[list[str]], *, header: bool = True) -> None:
    for r_i, row in enumerate(rows, start=1):
        for c_i, value in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=_cell_write_value(value))
            cell.font = _HEADER_FONT if header and r_i == 1 else _BODY_FONT
            if header and r_i == 1:
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
    if rows:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max(len(row) for row in rows))}{len(rows)}"


def _autosize(ws: Worksheet, rows: list[list[str]]) -> None:
    if not rows:
        return
    width = max(len(row) for row in rows)
    for c_i in range(1, width + 1):
        longest = 0
        for row in rows:
            if c_i - 1 < len(row):
                longest = max(longest, len(str(row[c_i - 1])))
        ws.column_dimensions[get_column_letter(c_i)].width = min(42, max(12, longest + 3))


def _cell_write_value(value: str):
    text = "" if value is None else str(value)
    if text.isdigit() and len(text) <= 10:
        try:
            return int(text)
        except ValueError:
            return text
    return text


def _sheet_title(name: str, index: int) -> str:
    title = (name or f"Sheet{index + 1}").strip() or f"Sheet{index + 1}"
    return title[:31]


def _apply_xlsx_op(workbook, tracker: LocationTracker, op: PatchOp) -> None:
    if op.op == "set_title":
        title = op.title or op.text
        if title is None:
            raise ArtifactPatchError("set_title requires 'title' or 'text'.")
        workbook.properties.title = title
        return
    if op.op == "insert_slide":
        raise ArtifactPatchError("insert_slide is only valid for ppt artifacts.")
    if op.op == "replace_shape_text":
        raise ArtifactPatchError("replace_shape_text is only valid for ppt artifacts.")
    if op.op == "set_cell":
        ws = _resolve_sheet(workbook, tracker, op)
        row, col = parse_cell_ref(op)
        ws.cell(row=row, column=col, value=cell_value(op))
        return
    if op.op == "replace_table":
        ws = _resolve_sheet(workbook, tracker, op, require_block=True)
        _replace_sheet_table(ws, require_table(op))
        return
    if op.op == "replace_text":
        block = tracker.find(op)
        ws = _sheet_named(workbook, block.location.sheet)
        coord = (block.location.cell_range or "A1").split(":", 1)[0]
        row, col = parse_cell_ref(PatchOp(op="set_cell", cell_range=coord))
        ws.cell(row=row, column=col, value=require_text(op))
        return
    if op.op == "insert_block":
        if op.block is None or not op.block.table:
            raise ArtifactPatchError("insert_block on excel requires a table block (new sheet).")
        name = str((op.block.extra or {}).get("sheet") or op.block.text or f"Sheet{len(workbook.sheetnames) + 1}")
        ws = workbook.create_sheet(title=_sheet_title(name, len(workbook.sheetnames)))
        _write_rows(ws, op.block.table)
        _autosize(ws, op.block.table)
        return
    if op.op == "delete_block":
        block = tracker.find(op)
        name = block.location.sheet
        if not name or name not in workbook.sheetnames:
            raise ArtifactPatchError(f"Sheet not found for block {block.block_id}.")
        if len(workbook.sheetnames) == 1:
            _replace_sheet_table(workbook[name], [])
        else:
            del workbook[name]
        tracker.remove(block)
        return
    raise ArtifactPatchError(f"Unknown patch op {op.op!r}.")


def _resolve_sheet(workbook, tracker: LocationTracker, op: PatchOp, *, require_block: bool = False):
    if op.block_id or op.location is not None:
        block = tracker.find(op)
        return _sheet_named(workbook, block.location.sheet)
    if op.sheet:
        return _sheet_named(workbook, op.sheet)
    if require_block:
        raise ArtifactPatchError(f"{op.op} requires block_id, location, or sheet.")
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise ArtifactPatchError("set_cell requires sheet, block_id, or a single-sheet workbook.")


def _sheet_named(workbook, name: str | None):
    if not name:
        raise ArtifactPatchError("Excel patch is missing a sheet name.")
    if name not in workbook.sheetnames:
        raise ArtifactPatchError(f"Sheet not found: {name}.")
    return workbook[name]


def _replace_sheet_table(ws: Worksheet, rows: list[list[str]]) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column and ws.max_column > 1:
        ws.delete_cols(1, ws.max_column)
    if rows:
        _write_rows(ws, rows)
        _autosize(ws, rows)
